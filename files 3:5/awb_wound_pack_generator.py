from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from io import BytesIO

from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image,
    KeepTogether,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from artifact_tool import SpreadsheetArtifact

BASE_DIR = Path("/mnt/data")
LOGO_PATH = BASE_DIR / "Teal_and_Navy_Logo_for_Healthcare_Brand-removebg-preview-1.webp"
PDF_PATH = BASE_DIR / "AWB_Wound_Documentation_Pack.pdf"
XLSX_PATH = BASE_DIR / "AWB_Wound_Tracking_Tool.xlsx"
PNG_LOGO_PATH = BASE_DIR / "AWB_logo_temp.png"

NAVY = colors.HexColor("#143A6F")
TEAL = colors.HexColor("#21869C")
LIGHT_TEAL = colors.HexColor("#E8F5F7")
PALE_BLUE = colors.HexColor("#EEF4FB")
SLATE = colors.HexColor("#4A5568")
LIGHT_GRAY = colors.HexColor("#D9E2EC")
SOFT_GRAY = colors.HexColor("#F7FAFC")
WHITE = colors.white

XL_NAVY = "143A6F"
XL_TEAL = "21869C"
XL_LIGHT_TEAL = "E8F5F7"
XL_PALE_BLUE = "EEF4FB"
XL_SOFT_GRAY = "F7FAFC"
XL_LIGHT_RED = "FDECEC"
XL_ORANGE = "FFF4E5"
XL_PURPLE = "5B3DBA"


def ensure_png_logo() -> Path:
    if PNG_LOGO_PATH.exists():
        return PNG_LOGO_PATH
    img = PILImage.open(LOGO_PATH).convert("RGBA")
    img.save(PNG_LOGO_PATH)
    return PNG_LOGO_PATH


def make_styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=25,
            leading=30,
            textColor=NAVY,
            alignment=TA_CENTER,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CoverSub",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=11,
            leading=15,
            textColor=SLATE,
            alignment=TA_CENTER,
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionHeader",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=18,
            textColor=WHITE,
            alignment=TA_LEFT,
            spaceBefore=6,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SubHeader",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            textColor=NAVY,
            spaceBefore=4,
            spaceAfter=4,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodySmall",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=SLATE,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodyBold",
            parent=styles["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            textColor=NAVY,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Tiny",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            textColor=SLATE,
        )
    )
    return styles


def section_band(text: str, styles):
    t = Table([[Paragraph(text, styles["SectionHeader"])]], colWidths=[7.0 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), NAVY),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("ROUNDEDCORNERS", [6, 6, 6, 6]),
            ]
        )
    )
    return t


def kv_table(rows, col_widths, bg=WHITE, styles=None, label_bg=PALE_BLUE):
    data = []
    for label, value in rows:
        data.append(
            [
                Paragraph(f"<b>{label}</b>", styles["BodySmall"]),
                Paragraph(value, styles["BodySmall"]),
            ]
        )
    t = Table(data, colWidths=col_widths, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), label_bg),
                ("BACKGROUND", (1, 0), (1, -1), bg),
                ("GRID", (0, 0), (-1, -1), 0.6, LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return t


def checklist_table(title, items, styles, prompt="Notes / dates / responsible party"):
    header = [
        Paragraph(f"<b>{title}</b>", styles["BodyBold"]),
        Paragraph("Y", styles["BodyBold"]),
        Paragraph("N", styles["BodyBold"]),
        Paragraph("N/A", styles["BodyBold"]),
        Paragraph(prompt, styles["BodyBold"]),
    ]
    rows = [header]
    for item in items:
        rows.append(
            [
                Paragraph(item, styles["BodySmall"]),
                Paragraph("[ ]", styles["BodySmall"]),
                Paragraph("[ ]", styles["BodySmall"]),
                Paragraph("[ ]", styles["BodySmall"]),
                Paragraph("", styles["BodySmall"]),
            ]
        )
    t = Table(rows, colWidths=[2.75 * inch, 0.45 * inch, 0.45 * inch, 0.55 * inch, 2.8 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), LIGHT_TEAL),
                ("GRID", (0, 0), (-1, -1), 0.6, LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (3, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return t


def narrative_box(title, prompts, styles, height_rows=4):
    header = [Paragraph(f"<b>{title}</b>", styles["BodyBold"])]
    lines = [Paragraph("<br/>".join([f"• {p}" for p in prompts]), styles["BodySmall"])]
    blank = [Paragraph("<br/>".join(["&nbsp;"] * height_rows), styles["BodySmall"])]
    t = Table([header, lines, blank], colWidths=[7.0 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PALE_BLUE),
                ("BACKGROUND", (0, 1), (-1, 1), WHITE),
                ("BACKGROUND", (0, 2), (-1, 2), SOFT_GRAY),
                ("BOX", (0, 0), (-1, -1), 0.8, LIGHT_GRAY),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, LIGHT_GRAY),
                ("LEFTPADDING", (0, 0), (-1, -1), 9),
                ("RIGHTPADDING", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return t


def conservative_care_table(styles):
    rows = [[
        Paragraph("<b>Week</b>", styles["BodyBold"]),
        Paragraph("<b>Date range</b>", styles["BodyBold"]),
        Paragraph("<b>Conservative care performed</b>", styles["BodyBold"]),
        Paragraph("<b>Objective response / wound trajectory</b>", styles["BodyBold"]),
    ]]
    for wk in range(1, 7):
        rows.append([
            Paragraph(str(wk), styles["BodySmall"]),
            Paragraph("", styles["BodySmall"]),
            Paragraph("", styles["BodySmall"]),
            Paragraph("", styles["BodySmall"]),
        ])
    t = Table(rows, colWidths=[0.55 * inch, 1.15 * inch, 2.65 * inch, 2.65 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), LIGHT_TEAL),
                ("GRID", (0, 0), (-1, -1), 0.6, LIGHT_GRAY),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return t


def footer(canvas, doc):
    canvas.saveState()
    w, h = letter
    canvas.setStrokeColor(LIGHT_GRAY)
    canvas.line(0.7 * inch, 0.62 * inch, w - 0.7 * inch, 0.62 * inch)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(SLATE)
    canvas.drawString(0.7 * inch, 0.42 * inch, "AWB Wound Documentation Pack • Standardized cross-setting documentation")
    canvas.drawRightString(w - 0.7 * inch, 0.42 * inch, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf():
    styles = make_styles()
    ensure_png_logo()

    doc = BaseDocTemplate(
        str(PDF_PATH),
        pagesize=letter,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.85 * inch,
        title="AWB Wound Documentation Pack",
        author="OpenAI",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=footer)])

    story = []

    logo = Image(str(PNG_LOGO_PATH), width=1.55 * inch, height=1.55 * inch)
    logo.hAlign = "CENTER"
    story.extend([
        Spacer(1, 0.35 * inch),
        logo,
        Spacer(1, 0.12 * inch),
        Paragraph("AWB Wound Documentation Pack", styles["CoverTitle"]),
        Paragraph(
            "Single standardized documentation language across provider clinic, SNF / NH / ALF, ASC, and ortho workflows.",
            styles["CoverSub"],
        ),
    ])

    cover_box = Table(
        [[Paragraph(
            "<b>Use this pack to standardize barrier screening, document at least four weeks of conservative care, support medical necessity, and align language across all care settings.</b>",
            styles["BodySmall"],
        )]],
        colWidths=[6.2 * inch],
    )
    cover_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), LIGHT_TEAL),
        ("BOX", (0, 0), (-1, -1), 0.8, TEAL),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.extend([cover_box, Spacer(1, 0.22 * inch)])

    story.append(section_band("Standardized Documentation Language", styles))
    story.append(Spacer(1, 0.1 * inch))
    story.append(
        narrative_box(
            "Core cross-setting statements",
            [
                "Barrier screen completed or updated today; diabetes control, vascular status, nutrition/labs, and smoking/alcohol counseling were reviewed and documented.",
                "The wound has received conservative management for at least four weeks, including wound assessment, local care, offloading/compression when indicated, and risk-factor optimization.",
                "Despite compliant conservative care, wound progression remains inadequate; escalation is medically necessary to promote closure, reduce complications, and preserve function.",
            ],
            styles,
            height_rows=3,
        )
    )

    story.append(PageBreak())

    story.extend([
        section_band("Encounter Header", styles),
        Spacer(1, 0.1 * inch),
        kv_table([
            ("Patient name", "______________________________________________"),
            ("DOB / MRN", "______________________________________________"),
            ("Encounter date", "______________________________________________"),
            ("Care setting", "Clinic   [ ]   SNF / NH / ALF   [ ]   ASC   [ ]   Ortho   [ ]"),
            ("Provider / facility", "______________________________________________"),
            ("Wound type / location", "______________________________________________"),
        ], [1.7 * inch, 5.3 * inch], styles=styles),
        Spacer(1, 0.14 * inch),
        narrative_box(
            "Quick workflow checkpoints",
            [
                "Barrier screen current?",
                "4+ weeks of conservative care documented?",
                "Consent, goals, and failure narrative ready for submission?",
            ],
            styles,
            height_rows=4,
        ),
    ])

    story.append(PageBreak())

    story.extend([
        section_band("Barrier Screening", styles),
        Spacer(1, 0.1 * inch),
        Paragraph("Complete or update each barrier domain at intake and at clinically relevant follow-up visits.", styles["BodySmall"]),
        Spacer(1, 0.1 * inch),
        checklist_table("A1c and diabetes control", [
            "Diabetes / prediabetes status reviewed",
            "Most recent A1c and collection date documented",
            "Hyperglycemia management plan or referral documented when indicated",
            "Medication adherence / glucose monitoring barriers addressed",
        ], styles),
        Spacer(1, 0.12 * inch),
        checklist_table("Vascular tests", [
            "Pedal pulses / perfusion findings documented",
            "ABI / TBI / Doppler / vascular study reviewed or ordered when indicated",
            "Venous disease / edema findings addressed",
            "Compression or offloading plan linked to vascular findings",
        ], styles),
        Spacer(1, 0.12 * inch),
        checklist_table("Nutrition and labs", [
            "Nutrition risk or poor intake screened",
            "Relevant labs reviewed (example: CBC, CMP, protein markers, renal status) as clinically appropriate",
            "Protein / calorie support or dietitian referral documented when indicated",
            "Hydration or social barriers affecting healing addressed",
        ], styles),
        Spacer(1, 0.12 * inch),
        checklist_table("Smoking / alcohol counseling", [
            "Tobacco use assessed",
            "Alcohol / substance use assessed",
            "Counseling delivered and patient response documented",
            "Referral / quit resources offered when indicated",
        ], styles),
    ])

    story.append(PageBreak())

    story.extend([
        section_band("Medical Necessity Scaffold", styles),
        Spacer(1, 0.1 * inch),
        checklist_table("Required elements for escalation / advanced intervention", [
            "Conservative care documented for at least 4 weeks",
            "Informed consent / shared decision-making documented",
            "Treatment goals are measurable and tied to function / closure / complication reduction",
            "Failed conservative care narrative explains why escalation is needed now",
        ], styles, prompt="Source note / date / evidence"),
        Spacer(1, 0.14 * inch),
        Paragraph("Conservative care timeline", styles["SubHeader"]),
        conservative_care_table(styles),
        Spacer(1, 0.14 * inch),
        narrative_box(
            "Failed conservative care narrative prompt",
            [
                "What standard wound care was provided over the prior 4+ weeks?",
                "What objective evidence shows inadequate progress (size, depth, tissue quality, drainage, pain, infection risk, function)?",
                "Why is the requested intervention expected to improve healing, safety, or limb/function preservation?",
            ],
            styles,
            height_rows=6,
        ),
    ])

    story.append(PageBreak())

    story.extend([
        section_band("Consent, Goals, and Attestation", styles),
        Spacer(1, 0.1 * inch),
        checklist_table("Consent", [
            "Risks, benefits, and alternatives discussed",
            "Patient / proxy questions answered",
            "Consent obtained and placed in chart",
        ], styles, prompt="Date / witness / document location"),
        Spacer(1, 0.12 * inch),
        checklist_table("Goals", [
            "Reduce wound surface area over defined interval",
            "Control drainage / odor / bioburden",
            "Prepare wound bed for closure or grafting",
            "Support mobility, limb preservation, or pain reduction",
        ], styles, prompt="Target metric / timeframe"),
        Spacer(1, 0.14 * inch),
        narrative_box(
            "Final medical necessity statement",
            [
                "Requested service / product / procedure:",
                "Clinical rationale linked to wound status and failed conservative care:",
                "Why delay would increase risk, cost, or morbidity:",
            ],
            styles,
            height_rows=7,
        ),
        Spacer(1, 0.18 * inch),
        kv_table([
            ("Provider signature", "____________________________________________________________"),
            ("Credentials", "____________________________________________________________"),
            ("Date", "____________________________________________________________"),
        ], [1.6 * inch, 5.4 * inch], styles=styles, label_bg=LIGHT_TEAL),
    ])

    doc.build(story)


def apply_cell_style(cell, *, fill=None, font=None, align=None, border=None, number_format=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def build_workbook():
    ensure_png_logo()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracker"

    thin_gray = Side(style="thin", color="D9E2EC")
    medium_navy = Side(style="medium", color=XL_NAVY)
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    fill_navy = PatternFill("solid", fgColor=XL_NAVY)
    fill_teal = PatternFill("solid", fgColor=XL_TEAL)
    fill_light_teal = PatternFill("solid", fgColor=XL_LIGHT_TEAL)
    fill_pale_blue = PatternFill("solid", fgColor=XL_PALE_BLUE)
    fill_soft_gray = PatternFill("solid", fgColor=XL_SOFT_GRAY)
    fill_red = PatternFill("solid", fgColor=XL_LIGHT_RED)
    fill_orange = PatternFill("solid", fgColor=XL_ORANGE)

    font_white_bold = Font(color="FFFFFF", bold=True, size=11)
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_bold = Font(color=XL_NAVY, bold=True)
    font_input = Font(color="0000FF")
    font_formula = Font(color="000000")
    font_static = Font(color="666666")
    font_green = Font(color="008000")
    font_purple = Font(color=XL_PURPLE, bold=True)
    font_small = Font(size=9)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Tracker title block
    ws.merge_cells("A1:V1")
    ws["A1"] = "AWB Wound Tracking Tool"
    apply_cell_style(ws["A1"], fill=fill_navy, font=Font(color="FFFFFF", bold=True, size=14), align=left)
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:V2")
    ws["A2"] = "Use one row per wound episode. Blue text = user input. Black text = formulas. Orange cells indicate review work."
    apply_cell_style(ws["A2"], fill=fill_light_teal, font=Font(color=XL_NAVY, italic=True, size=10), align=left, border=border)

    headers = [
        "Episode ID", "Patient Name", "DOB", "MRN", "Care Setting", "Provider", "Episode Start",
        "Wound Type", "Location", "A1c Status", "Vascular Status", "Nutrition/Labs Status",
        "Smoking/Alcohol Counseling", "Barrier Screen Status", "Conservative Care Start",
        "Weeks Conservative Care", "Consent", "Goals", "Failed Conservative Care Narrative",
        "Medical Necessity Ready", "Next Action", "Review Date", "Notes"
    ]
    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, header)
        apply_cell_style(cell, fill=fill_teal, font=font_header, align=center, border=border)

    widths = {
        1: 12, 2: 20, 3: 11, 4: 12, 5: 14, 6: 18, 7: 13, 8: 15, 9: 15,
        10: 12, 11: 12, 12: 14, 13: 14, 14: 14, 15: 16, 16: 12, 17: 11,
        18: 11, 19: 16, 20: 16, 21: 18, 22: 12, 23: 22
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in range(5, 55):
        ws.row_dimensions[row].height = 24
        for col in range(1, 24):
            cell = ws.cell(row, col)
            apply_cell_style(cell, fill=fill_soft_gray if row % 2 == 0 else PatternFill(fill_type=None), align=left, border=border)

        # formulas
        ws.cell(row, 14).value = f'=IF(COUNTA(J{row}:M{row})=0,"",IF(COUNTIF(J{row}:M{row},"Complete")=4,"Complete","Pending"))'
        ws.cell(row, 16).value = f'=IF(O{row}="","",INT((TODAY()-O{row})/7))'
        ws.cell(row, 20).value = (
            f'=IF(AND(N{row}="Complete",P{row}>=4,Q{row}="Yes",R{row}="Yes",S{row}="Yes"),'
            '"READY","INCOMPLETE")'
        )
        for formula_col in [14, 16, 20]:
            apply_cell_style(ws.cell(row, formula_col), font=font_formula)

    # Format user input columns in blue
    input_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 17, 18, 19, 21, 22, 23]
    for row in range(5, 55):
        for col in input_cols:
            ws.cell(row, col).font = font_input
        for col in [3, 7, 15, 22]:
            ws.cell(row, col).number_format = "mm/dd/yyyy"
        ws.cell(row, 16).number_format = '0" wks"'

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = "A4:W54"

    # Conditional formatting
    green_fill = PatternFill("solid", fgColor="E9F7EF")
    ws.conditional_formatting.add("T5:T54", FormulaRule(formula=["$T5=\"READY\""], fill=green_fill))
    ws.conditional_formatting.add("T5:T54", FormulaRule(formula=["$T5=\"INCOMPLETE\""], fill=fill_red))
    ws.conditional_formatting.add("N5:N54", FormulaRule(formula=["$N5=\"Pending\""], fill=fill_orange))

    # Instructions sheet
    instr = wb.create_sheet("Instructions")
    instr.sheet_view.showGridLines = False
    instr.merge_cells("A1:H1")
    instr["A1"] = "AWB Wound Documentation Workflow"
    apply_cell_style(instr["A1"], fill=fill_navy, font=font_white_bold, align=left)
    instr.column_dimensions["A"].width = 22
    instr.column_dimensions["B"].width = 22
    instr.column_dimensions["C"].width = 22
    instr.column_dimensions["D"].width = 22
    instr.column_dimensions["E"].width = 18
    instr.column_dimensions["F"].width = 18
    instr.column_dimensions["G"].width = 18
    instr.column_dimensions["H"].width = 18
    blocks = [
        (3, "1. Create an episode", "Enter demographic and wound information in Tracker columns A–I."),
        (5, "2. Complete barrier screening", "Use Complete / Pending / Not Needed in columns J–M. Column N calculates overall status."),
        (7, "3. Track conservative care", "Enter the conservative care start date in column O. Column P calculates elapsed weeks."),
        (9, "4. Validate scaffold items", "Columns Q–S capture consent, goals, and failed conservative care narrative."),
        (11, "5. Ready for escalation", "Column T returns READY when barrier screen is complete and all required elements are satisfied."),
        (13, "Suggested narrative", "Despite 4+ weeks of conservative care, wound progress remains inadequate. Escalation is medically necessary to promote healing, reduce complications, and preserve function.")
    ]
    for r, h, b in blocks:
        instr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        instr.cell(r, 1).value = h
        apply_cell_style(instr.cell(r, 1), fill=fill_teal, font=font_white_bold, align=left, border=border)
        instr.row_dimensions[r].height = 22
        if r == 13:
            instr.merge_cells(start_row=r+1, start_column=1, end_row=r+2, end_column=6)
            instr.cell(r+1, 1).value = b
            apply_cell_style(instr.cell(r+1, 1), fill=fill_light_teal, font=font_static, align=left, border=border)
            instr.row_dimensions[r+1].height = 42
            instr.row_dimensions[r+2].height = 22
        else:
            instr.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=6)
            instr.cell(r+1, 1).value = b
            apply_cell_style(instr.cell(r+1, 1), fill=fill_light_teal, font=font_static, align=left, border=border)
            instr.row_dimensions[r+1].height = 24

    # Insert logo in instructions if possible
    try:
        logo = XLImage(str(PNG_LOGO_PATH))
        logo.width = 90
        logo.height = 90
        instr.add_image(logo, "G3")
    except Exception:
        pass

    # Lists sheet for validations
    lists = wb.create_sheet("Lists")
    list_values = {
        "A": ["Clinic", "SNF/NH/ALF", "ASC", "Ortho"],
        "B": ["Complete", "Pending", "Not Needed"],
        "C": ["Yes", "No"],
        "D": ["Order tests", "Close barriers", "Continue conservative care", "Prepare request", "Escalate now"],
    }
    for col, values in list_values.items():
        for idx, value in enumerate(values, start=1):
            lists[f"{col}{idx}"] = value
    lists.sheet_state = "hidden"

    dv_setting = DataValidation(type="list", formula1="=Lists!$A$1:$A$4", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="=Lists!$B$1:$B$3", allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1="=Lists!$C$1:$C$2", allow_blank=True)
    dv_next = DataValidation(type="list", formula1="=Lists!$D$1:$D$5", allow_blank=True)

    ws.add_data_validation(dv_setting)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_yesno)
    ws.add_data_validation(dv_next)

    for row in range(5, 55):
        dv_setting.add(ws.cell(row, 5))
        for col in [10, 11, 12, 13]:
            dv_status.add(ws.cell(row, col))
        for col in [17, 18, 19]:
            dv_yesno.add(ws.cell(row, col))
        dv_next.add(ws.cell(row, 21))

    # Comments / source prompts
    ws["J4"].comment = Comment("Use status language that matches the PDF pack: Complete, Pending, or Not Needed.", "OpenAI")
    ws["T4"].comment = Comment("READY requires Barrier Screen Status = Complete, Weeks Conservative Care >= 4, and Yes in Consent, Goals, and Failed Conservative Care Narrative.", "OpenAI")

    # Sample starter row
    sample_row = 5
    sample_values = {
        1: "EX-001",
        2: "Sample Patient",
        5: "Clinic",
        6: "Provider Name",
        8: "Diabetic foot ulcer",
        9: "Left plantar hallux",
        10: "Complete",
        11: "Pending",
        12: "Complete",
        13: "Complete",
        17: "Yes",
        18: "Yes",
        19: "No",
        21: "Close barriers",
        23: "Replace sample row with active patient data.",
    }
    for col, val in sample_values.items():
        ws.cell(sample_row, col).value = val
    ws.cell(sample_row, 3).value = date(1965, 1, 15)
    ws.cell(sample_row, 7).value = date.today()
    ws.cell(sample_row, 15).value = date.today() - timedelta(days=35)
    ws.cell(sample_row, 22).value = date.today()

    wb.save(XLSX_PATH)

    # Recalculate/cache formulas and render for QA
    artifact = SpreadsheetArtifact.load(str(XLSX_PATH))
    artifact.recalculate()
    artifact.export(str(XLSX_PATH), overwrite=True)
    artifact.render(str(BASE_DIR / "awb_tracker_render"))


def main():
    build_pdf()
    build_workbook()
    print(f"Created: {PDF_PATH}")
    print(f"Created: {XLSX_PATH}")


if __name__ == "__main__":
    main()
